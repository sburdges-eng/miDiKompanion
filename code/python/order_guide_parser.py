"""
Order Guide Parser
Parse vendor order guides from Sysco and Shamrock
"""

import openpyxl
from typing import List, Dict, Optional
import logging
from pathlib import Path
import json

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class OrderGuideParser:
    """Parse order guides from multiple vendors"""

    def __init__(self):
        self.vendors = {
            'sysco': SyscoParser(),
            'shamrock': ShamrockParser(),
            'combo': ComboParser(),
            'official_combined': OfficialCombinedParser()
        }

    def parse_file(self, file_path: str, vendor: str = None) -> Dict:
        """
        Parse an order guide file

        Args:
            file_path: Path to Excel file
            vendor: Vendor name ('sysco', 'shamrock', 'combo') or None to auto-detect

        Returns:
            Dict with parsed products
        """
        if vendor and vendor.lower() in self.vendors:
            parser = self.vendors[vendor.lower()]
            return parser.parse(file_path)

        # Auto-detect vendor from filename or sheet names
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = [name.lower() for name in wb.sheetnames]

        if 'combo ph' in sheet_names:
            logger.info("Auto-detected: Combo template")
            return self.vendors['combo'].parse(file_path)
        elif 'sysco ph' in sheet_names or any('sysco' in name for name in sheet_names):
            logger.info("Auto-detected: Sysco template")
            return self.vendors['sysco'].parse(file_path)
        elif 'sham ph' in sheet_names or any('sham' in name for name in sheet_names):
            logger.info("Auto-detected: Shamrock template")
            return self.vendors['shamrock'].parse(file_path)
        else:
            raise ValueError(f"Unable to detect vendor format. Sheet names: {wb.sheetnames}")

    def parse_all_vendors(self, file_path: str) -> Dict:
        """
        Parse all vendor sheets in a combined file

        Returns:
            Dict with products from each vendor
        """
        wb = openpyxl.load_workbook(file_path, data_only=True)
        results = {}

        # Parse SYSCO PH
        if 'SYSCO PH ' in wb.sheetnames or 'SYSCO PH' in wb.sheetnames:
            sheet_name = 'SYSCO PH ' if 'SYSCO PH ' in wb.sheetnames else 'SYSCO PH'
            logger.info(f"Parsing {sheet_name}")
            results['sysco'] = self.vendors['sysco'].parse_sheet(wb[sheet_name])

        # Parse SHAM PH (Shamrock)
        if 'SHAM PH' in wb.sheetnames:
            logger.info("Parsing SHAM PH")
            results['shamrock'] = self.vendors['shamrock'].parse_sheet(wb['SHAM PH'])

        # Parse COMBO PH
        if 'COMBO PH' in wb.sheetnames:
            logger.info("Parsing COMBO PH")
            results['combo'] = self.vendors['combo'].parse_sheet(wb['COMBO PH'])

        return results


class SyscoParser:
    """Parse Sysco order guide format"""

    def parse(self, file_path: str) -> Dict:
        """Parse Sysco file"""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        # Try both sheet name variations
        sheet_name = 'SYSCO PH ' if 'SYSCO PH ' in wb.sheetnames else 'SYSCO PH'
        ws = wb[sheet_name]
        return self.parse_sheet(ws)

    def parse_sheet(self, ws) -> Dict:
        """Parse Sysco worksheet"""
        products = []

        # Row 1: Header (H | L0601 | ...)
        # Row 2: Field names (F | SUPC | Case Qty | Split Qty | ...)
        # Row 3+: Products (P | product_code | ...)

        # Get column mapping from row 2
        headers = {}
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(2, col_idx).value
            if cell_value:
                headers[str(cell_value).strip()] = col_idx

        logger.info(f"Sysco headers: {list(headers.keys())[:15]}")

        # Parse products starting from row 3
        for row_idx in range(3, ws.max_row + 1):
            # Check if it's a product row (starts with 'P')
            row_type = ws.cell(row_idx, 1).value
            if row_type != 'P':
                continue

            product = self._extract_product(ws, row_idx, headers)
            if product:
                products.append(product)

        logger.info(f"Parsed {len(products)} Sysco products")
        return {
            'vendor': 'Sysco',
            'format': 'SYSCO PH',
            'total_products': len(products),
            'products': products
        }

    def _extract_product(self, ws, row_idx: int, headers: Dict) -> Optional[Dict]:
        """Extract product data from row"""
        try:
            supc = ws.cell(row_idx, headers.get('SUPC', 2)).value
            if not supc:
                return None

            product = {
                'vendor': 'Sysco',
                'product_code': str(supc),
                'pack': ws.cell(row_idx, headers.get('Pack', 8)).value,
                'size': ws.cell(row_idx, headers.get('Size', 9)).value,
                'unit': ws.cell(row_idx, headers.get('Unit', 10)).value,
                'brand': ws.cell(row_idx, headers.get('Brand', 11)).value,
                'mfr_number': ws.cell(row_idx, headers.get('Mfr #', 12)).value,
                'case_qty': ws.cell(row_idx, headers.get('Case Qty', 3)).value,
                'split_qty': ws.cell(row_idx, headers.get('Split Qty', 4)).value,
            }

            # Try to get description from additional columns
            if 'Description' in headers:
                product['description'] = ws.cell(row_idx, headers['Description']).value
            elif 'Item' in headers:
                product['description'] = ws.cell(row_idx, headers['Item']).value

            # Get pricing if available
            if 'Price' in headers:
                product['price'] = ws.cell(row_idx, headers['Price']).value

            return product
        except Exception as e:
            logger.warning(f"Error extracting product at row {row_idx}: {e}")
            return None


class ShamrockParser:
    """Parse Shamrock order guide format"""

    def parse(self, file_path: str) -> Dict:
        """Parse Shamrock file"""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb['SHAM PH']
        return self.parse_sheet(ws)

    def parse_sheet(self, ws) -> Dict:
        """Parse Shamrock worksheet"""
        products = []

        # Shamrock format:
        # Rows 1-4: Header/metadata
        # Row 5: Column headers (# | Product # | Description | | | Pack Size | Brand | LWP | Avg | Price | Unit)
        # Row 6+: Products

        # Get column mapping from row 5
        headers = {}
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(5, col_idx).value
            if cell_value:
                headers[str(cell_value).strip()] = col_idx

        logger.info(f"Shamrock headers: {list(headers.keys())}")

        # Parse products starting from row 6
        for row_idx in range(6, ws.max_row + 1):
            product = self._extract_product(ws, row_idx, headers)
            if product:
                products.append(product)

        logger.info(f"Parsed {len(products)} Shamrock products")
        return {
            'vendor': 'Shamrock',
            'format': 'SHAM PH',
            'total_products': len(products),
            'products': products
        }

    def _extract_product(self, ws, row_idx: int, headers: Dict) -> Optional[Dict]:
        """Extract product data from row"""
        try:
            product_num = ws.cell(row_idx, headers.get('Product #', 3)).value
            if not product_num:
                return None

            # Clean price (remove $ and convert to float)
            price_cell = ws.cell(row_idx, headers.get('Price', 11)).value
            price = None
            if price_cell:
                price_str = str(price_cell).replace('$', '').replace(',', '').strip()
                try:
                    price = float(price_str)
                except:
                    price = price_str

            product = {
                'vendor': 'Shamrock',
                'product_code': str(product_num),
                'description': ws.cell(row_idx, headers.get('Description', 4)).value,
                'pack_size': ws.cell(row_idx, headers.get('Pack Size', 7)).value,
                'brand': ws.cell(row_idx, headers.get('Brand', 8)).value,
                'lwp': ws.cell(row_idx, headers.get('LWP', 9)).value,  # Last Week's Purchase
                'avg': ws.cell(row_idx, headers.get('Avg', 10)).value,  # Average usage
                'price': price,
                'unit': ws.cell(row_idx, headers.get('Unit', 12)).value,
            }

            return product
        except Exception as e:
            logger.warning(f"Error extracting product at row {row_idx}: {e}")
            return None


class ComboParser:
    """Parse combined order guide format (Sysco + Shamrock + purchase history)"""

    def parse(self, file_path: str) -> Dict:
        """Parse Combo file"""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb['COMBO PH']
        return self.parse_sheet(ws)

    def parse_sheet(self, ws) -> Dict:
        """Parse Combo worksheet"""
        products = []

        # Combo format is similar to Sysco but with additional columns (purchase history)
        # Row 1: Header
        # Row 2: Field names
        # Row 3+: Products

        # Get column mapping from row 2
        headers = {}
        for col_idx in range(1, min(50, ws.max_column + 1)):  # Limit to first 50 columns for headers
            cell_value = ws.cell(2, col_idx).value
            if cell_value:
                headers[str(cell_value).strip()] = col_idx

        logger.info(f"Combo headers: {list(headers.keys())[:20]}")

        # Parse products starting from row 3
        for row_idx in range(3, ws.max_row + 1):
            # Check if it's a product row (starts with 'P')
            row_type = ws.cell(row_idx, 1).value
            if row_type != 'P':
                continue

            product = self._extract_product(ws, row_idx, headers)
            if product:
                products.append(product)

        logger.info(f"Parsed {len(products)} Combo products")
        return {
            'vendor': 'Combined',
            'format': 'COMBO PH',
            'total_products': len(products),
            'products': products
        }

    def _extract_product(self, ws, row_idx: int, headers: Dict) -> Optional[Dict]:
        """Extract product data from row"""
        try:
            supc = ws.cell(row_idx, headers.get('SUPC', 2)).value
            if not supc:
                return None

            product = {
                'vendor': 'Combined',
                'product_code': str(supc),
                'pack': ws.cell(row_idx, headers.get('Pack', 8)).value,
                'size': ws.cell(row_idx, headers.get('Size', 9)).value,
                'unit': ws.cell(row_idx, headers.get('Unit', 10)).value,
                'brand': ws.cell(row_idx, headers.get('Brand', 11)).value,
                'mfr_number': ws.cell(row_idx, headers.get('Mfr #', 12)).value,
                'case_qty': ws.cell(row_idx, headers.get('Case Qty', 3)).value,
                'split_qty': ws.cell(row_idx, headers.get('Split Qty', 4)).value,
            }

            # Try to get description
            if 'Description' in headers:
                product['description'] = ws.cell(row_idx, headers['Description']).value
            elif 'Item' in headers:
                product['description'] = ws.cell(row_idx, headers['Item']).value

            # Get pricing if available
            if 'Price' in headers:
                product['price'] = ws.cell(row_idx, headers['Price']).value

            # Extract purchase history columns if available (columns beyond standard fields)
            # This could include weekly/monthly purchase quantities
            product['purchase_history'] = {}
            for header_name, col_idx in headers.items():
                if col_idx > 20:  # Columns after standard fields are likely purchase history
                    value = ws.cell(row_idx, col_idx).value
                    if value:
                        product['purchase_history'][header_name] = value

            return product
        except Exception as e:
            logger.warning(f"Error extracting product at row {row_idx}: {e}")
            return None


def create_unified_catalog(sysco_products: List[Dict], shamrock_products: List[Dict]) -> Dict:
    """
    Create unified product catalog from multiple vendors

    Args:
        sysco_products: List of Sysco products
        shamrock_products: List of Shamrock products

    Returns:
        Unified catalog with cross-reference
    """
    catalog = {
        'products': [],
        'vendor_map': {},
        'total_unique_products': 0
    }

    # Add all products with vendor tracking
    product_id = 1

    for product in sysco_products:
        unified_product = {
            'id': f'CAT-{product_id:06d}',
            'vendors': ['Sysco'],
            'sysco_code': product.get('product_code'),
            'description': product.get('description'),
            'pack': product.get('pack'),
            'size': product.get('size'),
            'unit': product.get('unit'),
            'brand': product.get('brand'),
            'sysco_price': product.get('price'),
        }
        catalog['products'].append(unified_product)
        catalog['vendor_map'][f"SYSCO_{product.get('product_code')}"] = product_id
        product_id += 1

    for product in shamrock_products:
        unified_product = {
            'id': f'CAT-{product_id:06d}',
            'vendors': ['Shamrock'],
            'shamrock_code': product.get('product_code'),
            'description': product.get('description'),
            'pack_size': product.get('pack_size'),
            'unit': product.get('unit'),
            'brand': product.get('brand'),
            'shamrock_price': product.get('price'),
        }
        catalog['products'].append(unified_product)
        catalog['vendor_map'][f"SHAM_{product.get('product_code')}"] = product_id
        product_id += 1

    catalog['total_unique_products'] = len(catalog['products'])

    return catalog


class OfficialCombinedParser:
    """
    Parser for Official Combined Order Guide Format
    This format has side-by-side vendor comparison with correct product numbers

    Format (Row 2 headers):
        Col 4: PRODUCT # (Shamrock)
        Col 6: DESCRIPTION (Shamrock)
        Col 12: PRODUCT # (Sysco)
        Col 14: DESCRIPTION (Sysco)

    Each row represents a matched pair of products from both vendors
    """

    def parse(self, file_path: str) -> Dict:
        """Parse official combined order guide"""
        logger.info(f"Parsing Official Combined Order Guide: {file_path}")

        wb = openpyxl.load_workbook(file_path, data_only=True)

        # Look for the combined guide sheet
        sheet_name = None
        for name in wb.sheetnames:
            if 'COMBINED' in name.upper() or 'GUIDE' in name.upper():
                sheet_name = name
                break

        if not sheet_name:
            # Default to first sheet
            sheet_name = wb.sheetnames[0]

        ws = wb[sheet_name]
        logger.info(f"Using sheet: {sheet_name}")

        sysco_products = []
        shamrock_products = []
        matched_pairs = []

        # Parse rows starting from row 3 (after headers in rows 1-2)
        for row_idx in range(3, ws.max_row + 1):
            # Shamrock columns
            sham_product_code = ws.cell(row_idx, 4).value  # Col 4
            sham_par = ws.cell(row_idx, 5).value           # Col 5
            sham_description = ws.cell(row_idx, 6).value   # Col 6
            sham_list = ws.cell(row_idx, 7).value          # Col 7
            sham_quantity = ws.cell(row_idx, 9).value      # Col 9
            sham_amount = ws.cell(row_idx, 10).value       # Col 10
            sham_size = ws.cell(row_idx, 11).value         # Col 11

            # Sysco columns
            sysco_product_code = ws.cell(row_idx, 12).value  # Col 12
            sysco_par = ws.cell(row_idx, 13).value           # Col 13
            sysco_description = ws.cell(row_idx, 14).value   # Col 14
            sysco_quantity = ws.cell(row_idx, 15).value      # Col 15
            sysco_amount = ws.cell(row_idx, 16).value        # Col 16
            sysco_size = ws.cell(row_idx, 17).value          # Col 17

            # Parse Shamrock product if present
            if sham_product_code and str(sham_product_code).strip():
                # Clean product code (remove leading zeros if numeric)
                product_code = str(sham_product_code).strip()

                shamrock_product = {
                    'vendor': 'Shamrock',
                    'product_code': product_code,
                    'description': str(sham_description).strip() if sham_description else '',
                    'par': sham_par,
                    'list_number': sham_list,
                    'quantity': sham_quantity,
                    'amount': sham_amount,
                    'size': str(sham_size).strip() if sham_size else '',
                    'pack_size': str(sham_size).strip() if sham_size else '',
                    'unit': 'CS',  # Default unit
                }
                shamrock_products.append(shamrock_product)

            # Parse Sysco product if present
            if sysco_product_code and str(sysco_product_code).strip():
                # Clean product code
                product_code = str(sysco_product_code).strip()

                # Skip if it's a formula (starts with =)
                if product_code.startswith('='):
                    continue

                sysco_product = {
                    'vendor': 'Sysco',
                    'product_code': product_code,
                    'description': str(sysco_description).strip() if sysco_description else '',
                    'par': sysco_par,
                    'quantity': sysco_quantity,
                    'amount': sysco_amount,
                    'size': str(sysco_size).strip() if sysco_size else '',
                    'pack': str(sysco_size).strip() if sysco_size else '',
                    'unit': 'CS',  # Default unit
                    'brand': '',  # Extract from description if possible
                }
                sysco_products.append(sysco_product)

            # If both vendors have products on this row, they are matched
            if (sham_product_code and str(sham_product_code).strip() and
                sysco_product_code and str(sysco_product_code).strip() and
                not str(sysco_product_code).startswith('=')):

                matched_pairs.append({
                    'shamrock_code': str(sham_product_code).strip(),
                    'shamrock_description': str(sham_description).strip() if sham_description else '',
                    'sysco_code': str(sysco_product_code).strip(),
                    'sysco_description': str(sysco_description).strip() if sysco_description else '',
                    'confidence': 1.0,  # Manual matches have 100% confidence
                    'source': 'official_guide'
                })

        logger.info(f"Parsed {len(shamrock_products)} Shamrock products")
        logger.info(f"Parsed {len(sysco_products)} Sysco products")
        logger.info(f"Found {len(matched_pairs)} pre-matched product pairs")

        return {
            'sysco': {
                'vendor': 'Sysco',
                'total_products': len(sysco_products),
                'products': sysco_products
            },
            'shamrock': {
                'vendor': 'Shamrock',
                'total_products': len(shamrock_products),
                'products': shamrock_products
            },
            'matched_pairs': matched_pairs,
            'format': 'official_combined',
            'source_file': file_path
        }


if __name__ == "__main__":
    # Test the parser
    file_path = "/Users/seanburdges/Downloads/COMBO PH NO TOUCH (1).xlsx"

    parser = OrderGuideParser()

    print("\n" + "=" * 80)
    print("ORDER GUIDE PARSER - TEST")
    print("=" * 80)

    # Parse all vendors
    results = parser.parse_all_vendors(file_path)

    print(f"\nVendors found: {list(results.keys())}")

    for vendor, data in results.items():
        print(f"\n{vendor.upper()}:")
        print(f"  Total products: {data['total_products']}")
        if data['products']:
            print(f"  Sample product: {data['products'][0]}")

    # Create unified catalog
    if 'sysco' in results and 'shamrock' in results:
        print("\n" + "=" * 80)
        print("CREATING UNIFIED CATALOG")
        print("=" * 80)

        catalog = create_unified_catalog(
            results['sysco']['products'],
            results['shamrock']['products']
        )

        print(f"\nTotal unique products: {catalog['total_unique_products']}")
        print(f"Sysco products: {len(results['sysco']['products'])}")
        print(f"Shamrock products: {len(results['shamrock']['products'])}")

        # Save catalog
        with open('unified_catalog.json', 'w') as f:
            json.dump(catalog, f, indent=2)
        print("\n✓ Saved unified catalog to unified_catalog.json")
