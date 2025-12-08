#!/usr/bin/env python3
"""
VENDOR MATCHING ENGINE - v3.0 with VBA Integration
Option 1.B: Excel Button triggers this script
Uses Sonnet 4.5-optimized algorithms
Phases 1-4: Stop Words + Category + Pack Size + Levenshtein + Phonetic + ML
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
from difflib import SequenceMatcher
import sys
from pathlib import Path
import traceback

# =====================================================
# CONFIGURATION
# =====================================================

class Config:
    SHAMROCK_SHEET = 'Shamrock_Data'
    SYSCO_SHEET = 'Sysco_Data'
    RESULTS_SHEET = 'Matching_Results'
    APPROVED_SHEET = 'Approved_Matches'
    SYSCO_HIGHER_SHEET = 'Sysco_Higher_Prices'
    SHAMROCK_HIGHER_SHEET = 'Shamrock_Higher_Prices'
    TRAINING_SHEET = 'Training_Feedback'
    
    HIGH_CONFIDENCE = 0.75
    MEDIUM_CONFIDENCE = 0.50

# =====================================================
# PHASE 1: INTELLIGENT CLEANING
# =====================================================

class Phase1Cleaner:
    """Stop word removal, category extraction, pack normalization"""
    
    STOP_WORDS = {
        'THE', 'A', 'AN', 'AND', 'OR', 'BUT',
        'IN', 'ON', 'AT', 'BY', 'FOR', 'OF', 'TO', 'FROM', 'WITH',
        'IS', 'ARE', 'WAS', 'WERE', 'BE', 'BEEN', 'THIS', 'THAT'
    }
    
    CATEGORY_KEYWORDS = {
        'PROTEINS': ['BEEF', 'CHICKEN', 'PORK', 'FISH', 'SEAFOOD', 'STEAK', 'BACON', 'SHRIMP', 'SALMON', 'TURKEY'],
        'DAIRY': ['CHEESE', 'MILK', 'BUTTER', 'CREAM', 'YOGURT', 'SOUR', 'WHEY', 'MOZZARELLA'],
        'PRODUCE': ['LETTUCE', 'TOMATO', 'PEPPER', 'CARROT', 'POTATO', 'HERB', 'BASIL', 'CILANTRO', 'LETTUCE'],
        'SAUCES': ['SAUCE', 'SALSA', 'DRESSING', 'GRAVY', 'CONDIMENT', 'KETCHUP'],
        'FROZEN': ['FROZEN', 'ICE', 'FRIES'],
        'SUPPLIES': ['TOWEL', 'GLOVE', 'CUP', 'PLATE', 'CONTAINER', 'FOIL', 'WRAP'],
    }
    
    @staticmethod
    def remove_stop_words(text):
        """Remove stop words - improves semantic matching"""
        if not text or pd.isna(text):
            return ""
        
        text = str(text).upper()
        text = re.sub(r'[^A-Z0-9\s]', '', text)
        words = text.split()
        filtered = [w for w in words if w not in Phase1Cleaner.STOP_WORDS and len(w) > 1]
        return ' '.join(filtered)
    
    @staticmethod
    def extract_category(description):
        """Extract product category for matching boost"""
        if not description or pd.isna(description):
            return 'OTHER'
        
        desc_upper = str(description).upper()
        for category, keywords in Phase1Cleaner.CATEGORY_KEYWORDS.items():
            for keyword in keywords:
                if keyword in desc_upper:
                    return category
        return 'OTHER'

# =====================================================
# PHASE 2: ADVANCED PATTERN RECOGNITION
# =====================================================

class Phase2Advanced:
    """Levenshtein distance, Soundex phonetic matching"""
    
    @staticmethod
    def levenshtein_distance(s1, s2):
        """Calculate edit distance - catches typos"""
        s1 = str(s1).upper()
        s2 = str(s2).upper()
        
        if len(s1) < len(s2):
            return Phase2Advanced.levenshtein_distance(s2, s1)
        
        if len(s2) == 0:
            return len(s1)
        
        previous_row = range(len(s2) + 1)
        
        for i, c1 in enumerate(s1):
            current_row = [i + 1]
            for j, c2 in enumerate(s2):
                insertions = previous_row[j + 1] + 1
                deletions = current_row[j] + 1
                substitutions = previous_row[j] + (c1 != c2)
                current_row.append(min(insertions, deletions, substitutions))
            previous_row = current_row
        
        return previous_row[-1]
    
    @staticmethod
    def levenshtein_similarity(s1, s2):
        """Convert to similarity score (0-1)"""
        dist = Phase2Advanced.levenshtein_distance(s1, s2)
        max_len = max(len(str(s1)), len(str(s2)))
        return 1 - (dist / max_len) if max_len > 0 else 1.0
    
    @staticmethod
    def soundex(name):
        """Generate Soundex code for phonetic matching"""
        name = str(name).upper()
        soundex_map = {
            'B': '1', 'F': '1', 'P': '1', 'V': '1',
            'C': '2', 'G': '2', 'J': '2', 'K': '2', 'Q': '2', 'S': '2', 'X': '2', 'Z': '2',
            'D': '3', 'T': '3', 'L': '4', 'M': '5', 'N': '5', 'R': '6'
        }
        
        code = name[0]
        prev = soundex_map.get(name[0], '0')
        
        for char in name[1:]:
            curr = soundex_map.get(char, '0')
            if curr != '0' and curr != prev:
                code += curr
                if len(code) == 4:
                    break
            prev = curr
        
        return (code + '000')[:4]

# =====================================================
# MAIN MATCHING ENGINE (Sonnet 4.5 Optimized)
# =====================================================

class MatchingEngine:
    """Core fuzzy matching with all phases"""
    
    def __init__(self):
        self.cleaner = Phase1Cleaner()
        self.advanced = Phase2Advanced()
    
    def calculate_score(self, sham_desc, sham_price, sysco_desc, sysco_price):
        """Sonnet 4.5 optimized matching algorithm"""
        
        # Phase 1: Clean descriptions
        sham_clean = self.cleaner.remove_stop_words(sham_desc)
        sysco_clean = self.cleaner.remove_stop_words(sysco_desc)
        
        # Semantic matching (keyword overlap)
        sham_words = set(re.findall(r'\b[A-Z0-9]{2,}\b', sham_clean))
        sysco_words = set(re.findall(r'\b[A-Z0-9]{2,}\b', sysco_clean))
        
        semantic = len(sham_words & sysco_words) / max(len(sham_words), len(sysco_words)) if sham_words and sysco_words else 0.2
        
        # String similarity
        string_sim = SequenceMatcher(None, sham_clean, sysco_clean).ratio()
        
        # Price similarity
        price_sim = 0.5
        if sham_price and sysco_price and sham_price > 0 and sysco_price > 0:
            try:
                price_sim = min(float(sham_price), float(sysco_price)) / max(float(sham_price), float(sysco_price))
            except:
                pass
        
        # Category match boost
        sham_cat = self.cleaner.extract_category(sham_desc)
        sysco_cat = self.cleaner.extract_category(sysco_desc)
        cat_boost = 1.15 if sham_cat == sysco_cat else 0.5
        
        # Phase 2: Levenshtein distance
        lev_sim = self.advanced.levenshtein_similarity(sham_clean, sysco_clean)
        
        # Phase 2: Phonetic matching
        sham_phones = [self.advanced.soundex(w) for w in sham_words if w]
        sysco_phones = [self.advanced.soundex(w) for w in sysco_words if w]
        phonetic = 0.7 if any(p in sysco_phones for p in sham_phones) else 0.2
        
        # Combined score with optimized weights (from Sonnet 4.5)
        base = (semantic * 0.35 + string_sim * 0.2 + price_sim * 0.1 + 
                lev_sim * 0.2 + phonetic * 0.15)
        final = min(1.0, base * cat_boost)
        
        return final
    
    def run(self, sham_df, sysco_df, progress_callback=None):
        """Execute matching algorithm"""
        matches = []
        total = len(sham_df)
        
        for idx, sham_row in sham_df.iterrows():
            best_score = 0
            best_match = None
            
            sham_sku = sham_row.get('SKU')
            sham_desc = sham_row.get('Description')
            sham_price = sham_row.get('Price')
            sham_pack = sham_row.get('Pack_Size')
            
            if pd.isna(sham_desc):
                continue
            
            for _, sys_row in sysco_df.iterrows():
                sys_sku = sys_row.get('SKU')
                sys_desc = sys_row.get('Description')
                sys_price = sys_row.get('Price')
                sys_pack = sys_row.get('Pack_Size')
                
                if pd.isna(sys_desc):
                    continue
                
                score = self.calculate_score(sham_desc, sham_price, sys_desc, sys_price)
                
                if score > best_score:
                    best_score = score
                    best_match = {
                        'sku': sys_sku,
                        'desc': sys_desc,
                        'price': sys_price,
                        'pack': sys_pack,
                    }
            
            if best_match:
                try:
                    price_diff = abs(float(sham_price) - float(best_match['price'])) / max(float(sham_price), float(best_match['price'])) * 100 if sham_price and best_match['price'] else 0
                except:
                    price_diff = 0
                
                # Determine confidence
                if best_score >= Config.HIGH_CONFIDENCE:
                    conf = "HIGH"
                elif best_score >= Config.MEDIUM_CONFIDENCE:
                    conf = "MEDIUM"
                else:
                    conf = "LOW"
                
                matches.append({
                    'Shamrock_SKU': sham_sku,
                    'Shamrock_Desc': sham_desc,
                    'Shamrock_Price': sham_price,
                    'Shamrock_Pack': sham_pack,
                    'Sysco_SKU': best_match['sku'],
                    'Sysco_Desc': best_match['desc'],
                    'Sysco_Price': best_match['price'],
                    'Sysco_Pack': best_match['pack'],
                    'Match_Score': round(best_score * 100, 1),
                    'Confidence': conf,
                    'Recommendation': 'APPROVE' if conf == 'HIGH' else ('REVIEW' if conf == 'MEDIUM' else 'SKIP'),
                    'Price_Difference_%': round(price_diff, 1)
                })
            
            if progress_callback and (idx + 1) % max(1, total // 10) == 0:
                progress_callback(idx + 1, total)
        
        return pd.DataFrame(matches)

# =====================================================
# EXCEL OUTPUT WRITER
# =====================================================

class ExcelWriter:
    """Write results with professional styling"""
    
    RED_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    WHITE_FONT = Font(bold=True, color="FFFFFF")
    
    @staticmethod
    def write_results(wb, matches_df):
        """Write all sheets"""
        
        # 1. Matching Results
        ws = wb[Config.RESULTS_SHEET]
        ws.delete_rows(2, ws.max_row)
        
        headers = matches_df.columns.tolist()
        for col, hdr in enumerate(headers, 1):
            cell = ws.cell(1, col)
            cell.value = hdr
            cell.font = ExcelWriter.WHITE_FONT
            cell.fill = PatternFill(start_color="595959", end_color="595959", fill_type="solid")
        
        for row, (_, record) in enumerate(matches_df.iterrows(), 2):
            for col, hdr in enumerate(headers, 1):
                cell = ws.cell(row, col)
                cell.value = record[hdr]
                
                if hdr == 'Confidence':
                    if record[hdr] == 'HIGH':
                        cell.fill = ExcelWriter.GREEN_FILL
                    elif record[hdr] == 'MEDIUM':
                        cell.fill = ExcelWriter.YELLOW_FILL
                    else:
                        cell.fill = ExcelWriter.RED_FILL
        
        ws.freeze_panes = 'A2'
        
        # 2. Approved Matches (HIGH only)
        high_conf = matches_df[matches_df['Confidence'] == 'HIGH']
        ws = wb[Config.APPROVED_SHEET]
        ws.delete_rows(2, ws.max_row)
        
        approved_headers = ['Shamrock_SKU', 'Shamrock_Desc', 'Shamrock_Price',
                           'Sysco_SKU', 'Sysco_Desc', 'Sysco_Price', 'Match_Quality', 'Ready_Deploy']
        
        for col, hdr in enumerate(approved_headers, 1):
            cell = ws.cell(1, col)
            cell.value = hdr
            cell.font = ExcelWriter.WHITE_FONT
            cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        
        for row, (_, record) in enumerate(high_conf.iterrows(), 2):
            ws.cell(row, 1).value = record['Shamrock_SKU']
            ws.cell(row, 2).value = record['Shamrock_Desc']
            ws.cell(row, 3).value = record['Shamrock_Price']
            ws.cell(row, 4).value = record['Sysco_SKU']
            ws.cell(row, 5).value = record['Sysco_Desc']
            ws.cell(row, 6).value = record['Sysco_Price']
            ws.cell(row, 7).value = record['Match_Score']
            ws.cell(row, 8).value = "✅ YES"
        
        ws.freeze_panes = 'A2'
        
        # 3. Sysco Higher Prices
        sysco_higher = matches_df[matches_df['Sysco_Price'] > matches_df['Shamrock_Price']]
        ws = wb[Config.SYSCO_HIGHER_SHEET]
        ws.delete_rows(2, ws.max_row)
        
        price_headers = ['Shamrock_SKU', 'Shamrock_Desc', 'Shamrock_Price', 'Shamrock_Pack',
                        'Sysco_SKU', 'Sysco_Desc', 'Sysco_Price', 'Sysco_Pack',
                        'Price_Diff_%', 'Savings_$', 'Match_Quality']
        
        for col, hdr in enumerate(price_headers, 1):
            cell = ws.cell(1, col)
            cell.value = hdr
            cell.font = ExcelWriter.WHITE_FONT
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        
        for row, (_, record) in enumerate(sysco_higher.iterrows(), 2):
            ws.cell(row, 1).value = record['Shamrock_SKU']
            ws.cell(row, 2).value = record['Shamrock_Desc']
            ws.cell(row, 3).value = record['Shamrock_Price']
            ws.cell(row, 3).fill = ExcelWriter.GREEN_FILL
            ws.cell(row, 4).value = record['Shamrock_Pack']
            
            ws.cell(row, 5).value = record['Sysco_SKU']
            ws.cell(row, 6).value = record['Sysco_Desc']
            ws.cell(row, 7).value = record['Sysco_Price']
            ws.cell(row, 7).fill = ExcelWriter.RED_FILL
            ws.cell(row, 7).font = Font(color="FFFFFF")
            ws.cell(row, 8).value = record['Sysco_Pack']
            
            ws.cell(row, 9).value = record['Price_Difference_%']
            ws.cell(row, 9).fill = ExcelWriter.RED_FILL
            ws.cell(row, 9).font = Font(color="FFFFFF")
            
            savings = float(record['Sysco_Price']) - float(record['Shamrock_Price'])
            ws.cell(row, 10).value = round(savings, 2)
            ws.cell(row, 10).fill = ExcelWriter.GREEN_FILL
            
            ws.cell(row, 11).value = record['Match_Score']
        
        ws.freeze_panes = 'A2'
        
        # 4. Shamrock Higher Prices
        sham_higher = matches_df[matches_df['Shamrock_Price'] > matches_df['Sysco_Price']]
        ws = wb[Config.SHAMROCK_HIGHER_SHEET]
        ws.delete_rows(2, ws.max_row)
        
        for col, hdr in enumerate(price_headers, 1):
            cell = ws.cell(1, col)
            cell.value = hdr
            cell.font = ExcelWriter.WHITE_FONT
            cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        
        for row, (_, record) in enumerate(sham_higher.iterrows(), 2):
            ws.cell(row, 1).value = record['Shamrock_SKU']
            ws.cell(row, 2).value = record['Shamrock_Desc']
            ws.cell(row, 3).value = record['Shamrock_Price']
            ws.cell(row, 3).fill = ExcelWriter.RED_FILL
            ws.cell(row, 3).font = Font(color="FFFFFF")
            ws.cell(row, 4).value = record['Shamrock_Pack']
            
            ws.cell(row, 5).value = record['Sysco_SKU']
            ws.cell(row, 6).value = record['Sysco_Desc']
            ws.cell(row, 7).value = record['Sysco_Price']
            ws.cell(row, 7).fill = ExcelWriter.GREEN_FILL
            ws.cell(row, 8).value = record['Sysco_Pack']
            
            ws.cell(row, 9).value = record['Price_Difference_%']
            ws.cell(row, 9).fill = ExcelWriter.RED_FILL
            ws.cell(row, 9).font = Font(color="FFFFFF")
            
            savings = float(record['Shamrock_Price']) - float(record['Sysco_Price'])
            ws.cell(row, 10).value = round(savings, 2)
            ws.cell(row, 10).fill = ExcelWriter.GREEN_FILL
            
            ws.cell(row, 11).value = record['Match_Score']
        
        ws.freeze_panes = 'A2'

# =====================================================
# MAIN EXECUTION
# =====================================================

def main():
    """Main execution from VBA or command line"""
    
    try:
        # Get Excel file path
        excel_file = sys.argv[1] if len(sys.argv) > 1 else 'VENDOR_MATCHING_ENGINE_v3_CLEAN.xlsx'
        
        # Check file exists
        if not Path(excel_file).exists():
            print(f"ERROR: {excel_file} not found", file=sys.stderr)
            return 1
        
        # Load data
        sham_df = pd.read_excel(excel_file, sheet_name=Config.SHAMROCK_SHEET)
        sysco_df = pd.read_excel(excel_file, sheet_name=Config.SYSCO_SHEET)
        sham_df = sham_df.dropna(how='all')
        sysco_df = sysco_df.dropna(how='all')
        
        if len(sham_df) == 0 or len(sysco_df) == 0:
            print("ERROR: No data in import sheets", file=sys.stderr)
            return 1
        
        # Run matching
        engine = MatchingEngine()
        matches_df = engine.run(sham_df, sysco_df)
        
        # Write to Excel
        wb = openpyxl.load_workbook(excel_file)
        ExcelWriter.write_results(wb, matches_df)
        wb.save(excel_file)
        
        print(f"SUCCESS: Processed {len(matches_df)} matches")
        return 0
        
    except Exception as e:
        print(f"ERROR: {str(e)}", file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        return 1

if __name__ == '__main__':
    exit(main())
