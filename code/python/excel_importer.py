"""
Excel Data Importer
Import data from Excel (.xlsx, .xls) files into the application
"""

import openpyxl
from openpyxl import load_workbook
from pathlib import Path
from typing import List, Dict, Optional, Any
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelImporter:
    """Import and process Excel files"""

    def __init__(self, data_directory: str = None):
        """
        Initialize Excel Importer

        Args:
            data_directory: Path to data directory
        """
        if data_directory is None:
            data_directory = Path(__file__).parent.parent / 'data'

        self.data_directory = Path(data_directory)
        self.imported_workbooks = {}
        self.imported_data = {}

    def list_excel_files(self, subdirectory: str = None) -> List[str]:
        """
        List all Excel files in directory

        Args:
            subdirectory: Optional subdirectory to search

        Returns:
            List of Excel file paths
        """
        search_path = self.data_directory
        if subdirectory:
            search_path = search_path / subdirectory

        if not search_path.exists():
            logger.warning(f"Directory does not exist: {search_path}")
            return []

        excel_files = list(search_path.glob('**/*.xlsx')) + list(search_path.glob('**/*.xls'))
        # Filter out temp files
        excel_files = [f for f in excel_files if not f.name.startswith('~$')]
        return [str(f.relative_to(self.data_directory)) for f in excel_files]

    def import_excel(self, filename: str, sheet_name: str = None,
                    has_header: bool = True, start_row: int = 1) -> Optional[List[Dict]]:
        """
        Import data from Excel file

        Args:
            filename: Excel filename
            sheet_name: Sheet name (uses first sheet if None)
            has_header: Whether first row contains headers
            start_row: Row number to start reading from (1-indexed)

        Returns:
            List of dictionaries representing rows
        """
        file_path = self.data_directory / filename

        if not file_path.exists():
            logger.error(f"File not found: {file_path}")
            return None

        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)

            # Get the sheet
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    logger.error(f"Sheet '{sheet_name}' not found in {filename}")
                    logger.info(f"Available sheets: {workbook.sheetnames}")
                    return None
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active

            # Read data
            data = []
            rows = list(sheet.iter_rows(min_row=start_row, values_only=True))

            if not rows:
                logger.warning(f"No data found in {filename}")
                return []

            if has_header:
                headers = rows[0]
                # Clean headers - remove None and convert to strings
                headers = [str(h).strip() if h is not None else f'Column_{i}'
                          for i, h in enumerate(headers)]

                for row in rows[1:]:
                    if any(cell is not None for cell in row):  # Skip empty rows
                        row_dict = {}
                        for header, cell in zip(headers, row):
                            row_dict[header] = cell
                        data.append(row_dict)
            else:
                for row in rows:
                    if any(cell is not None for cell in row):
                        row_dict = {f'col_{i}': cell for i, cell in enumerate(row)}
                        data.append(row_dict)

            workbook.close()

            key = f"{filename}:{sheet_name or 'default'}"
            self.imported_data[key] = data
            logger.info(f"Imported {len(data)} rows from {filename} (sheet: {sheet_name or sheet.title})")
            return data

        except Exception as e:
            logger.error(f"Error importing {filename}: {e}")
            return None

    def import_menu_bible(self, filename: str = "LARIAT MENU BIBLE PT.18 V3.xlsx") -> Optional[Dict]:
        """
        Import menu bible data

        Args:
            filename: Menu bible filename

        Returns:
            Dictionary with menu data organized by category
        """
        # Import from INDEX sheet which contains the menu overview
        data = self.import_excel(filename, sheet_name='INDEX')
        if not data:
            # Fallback to default sheet if INDEX doesn't exist
            data = self.import_excel(filename)
        if not data:
            return None

        menu_data = {
            'items': [],
            'categories': set(),
            'recipes': []
        }

        for row in data:
            item = {
                'name': row.get('Item') or row.get('item') or row.get('Recipe'),
                'category': row.get('Category') or row.get('category'),
                'cost': self._safe_float(row.get('Cost') or row.get('cost')),
                'price': self._safe_float(row.get('Price') or row.get('price')),
                'margin': self._safe_float(row.get('Margin') or row.get('margin')),
            }

            if item['name']:
                menu_data['items'].append(item)
                if item['category']:
                    menu_data['categories'].add(item['category'])

        menu_data['categories'] = list(menu_data['categories'])
        logger.info(f"Imported {len(menu_data['items'])} menu items in {len(menu_data['categories'])} categories")
        return menu_data

    def import_single_recipe(self, filename: str, recipe_name: str) -> Optional[Dict]:
        """
        Import a single recipe sheet with proper parsing

        Args:
            filename: Menu bible filename
            recipe_name: Name of the recipe sheet

        Returns:
            Dictionary with recipe details and ingredients
        """
        data = self.import_excel(filename, sheet_name=recipe_name)
        if not data or len(data) < 5:
            return None

        recipe = {
            'name': recipe_name,
            'category': None,
            'base_yield': None,
            'scaled_yield': None,
            'scale_factor': 1,
            'ingredients': []
        }

        # Row 0: Category info (Column_1 has category value)
        if len(data) > 0 and 'Column_1' in data[0]:
            recipe['category'] = data[0].get('Column_1')

        # Row 1: Yield info (Column_1 = base yield, Column_4 = scaled yield)
        if len(data) > 1:
            recipe['base_yield'] = data[1].get('Column_1')
            recipe['scaled_yield'] = data[1].get('Column_4')
            recipe['scale_factor'] = self._safe_float(data[0].get('Column_4'), 1)

        # Rows 4+: Ingredients (skip rows 2 and 3 which are headers)
        first_col = recipe_name  # The first column name is the recipe name
        for i in range(4, len(data)):
            row = data[i]

            # Column_1 = Ingredient name
            ingredient_name = row.get('Column_1')
            if not ingredient_name or str(ingredient_name).strip() == '':
                continue

            ingredient = {
                'name': str(ingredient_name).strip(),
                'base_qty': self._safe_float(row.get('Column_2')),
                'unit': row.get('Column_3'),
                'scaled_qty': self._safe_float(row.get('Column_4')),
            }

            recipe['ingredients'].append(ingredient)

        return recipe

    def import_menu_bible_recipes(self, filename: str = "LARIAT MENU BIBLE PT.18 V3.xlsx") -> Optional[Dict]:
        """
        Import all recipe sheets from menu bible

        Args:
            filename: Menu bible filename

        Returns:
            Dictionary with all recipe data from individual sheets
        """
        # Get all sheet names
        sheets = self.get_sheet_names(filename)
        if not sheets:
            return None

        all_recipes = {
            'recipes': [],
            'recipe_sheets': []
        }

        # Skip INDEX sheet, import all others as individual recipes
        for sheet_name in sheets:
            if sheet_name.upper() == 'INDEX':
                continue

            recipe = self.import_single_recipe(filename, sheet_name)
            if recipe:
                all_recipes['recipes'].append(recipe)
                all_recipes['recipe_sheets'].append(sheet_name)

        logger.info(f"Imported {len(all_recipes['recipes'])} recipe sheets from menu bible")
        return all_recipes

    def import_smart_costing(self, filename: str = "SMART COSTING 1/LARIAT_SMART_COSTING_COMPLETE_1.xlsx") -> Optional[Dict]:
        """
        Import smart costing data

        Args:
            filename: Smart costing filename

        Returns:
            Dictionary with costing data
        """
        # Try to import from multiple possible sheets
        # Based on actual file structure: SMART TEMPLATE, Ingredient Database, etc.
        sheets_to_try = ['SMART TEMPLATE', 'Ingredient Database', '🚀 START HERE',
                        'Costing', 'Recipes', 'Summary', 'Sheet1']

        for sheet_name in sheets_to_try:
            data = self.import_excel(filename, sheet_name=sheet_name)
            if data:
                costing_data = {
                    'recipes': [],
                    'total_cost': 0.0,
                    'sheet_name': sheet_name
                }

                for row in data:
                    recipe = {
                        'name': row.get('Recipe') or row.get('Item') or row.get('recipe_name'),
                        'servings': self._safe_int(row.get('Servings') or row.get('Yield')),
                        'total_cost': self._safe_float(row.get('Total Cost') or row.get('Cost')),
                        'cost_per_serving': self._safe_float(row.get('Cost Per Serving')),
                    }

                    if recipe['name']:
                        costing_data['recipes'].append(recipe)
                        costing_data['total_cost'] += recipe['total_cost']

                logger.info(f"Imported {len(costing_data['recipes'])} recipes from smart costing")
                return costing_data

        logger.error("Could not find valid sheet in smart costing file")
        return None

    def import_ingredient_database(self, filename: str = "SMART COSTING 1/LARIAT_SMART_COSTING_COMPLETE_1.xlsx") -> Optional[List[Dict]]:
        """
        Import ingredient database from smart costing file

        Args:
            filename: Smart costing filename

        Returns:
            List of ingredients with pricing and unit information
        """
        data = self.import_excel(filename, sheet_name='Ingredient Database')
        if not data:
            logger.warning("Could not import Ingredient Database sheet")
            return None

        ingredients = []
        # The first column has emoji and long name, but contains the ingredient names
        # Column_1 = CATEGORY, Column_2 = PURCHASE UNIT, Column_3 = PACKAGE SIZE, Column_4 = COST PER PKG
        ingredient_col = '🗄️ INGREDIENT DATABASE - MASTER PRICING & CONVERSIONS'

        for i, row in enumerate(data):
            # Skip header row (first row contains "INGREDIENT NAME")
            if i == 0 or row.get(ingredient_col) == 'INGREDIENT NAME':
                continue

            ingredient_name = row.get(ingredient_col)

            # Skip empty rows
            if not ingredient_name or str(ingredient_name).strip() == '':
                continue

            ingredient = {
                'name': str(ingredient_name).strip(),
                'category': row.get('Column_1'),
                'purchase_unit': row.get('Column_2'),
                'package_size': self._safe_float(row.get('Column_3')),
                'cost_per_package': self._safe_float(row.get('Column_4')),
                'cost_per_lb': self._safe_float(row.get('Column_5')),
                'cost_per_oz': self._safe_float(row.get('Column_6')),
            }

            ingredients.append(ingredient)

        logger.info(f"Imported {len(ingredients)} ingredients from database")
        return ingredients

    def import_vendor_matching(self, filename: str = "FINAL-OG-COMBO/VENDOR_MATCHING_PH_RESULTS.xlsx") -> Optional[Dict]:
        """
        Import vendor matching results

        Args:
            filename: Vendor matching filename

        Returns:
            Dictionary with vendor matching data
        """
        data = self.import_excel(filename)
        if not data:
            return None

        matching_data = {
            'matches': [],
            'vendors': set(),
            'unmatched': []
        }

        for row in data:
            match = {
                'product': row.get('Product') or row.get('Item'),
                'vendor1': row.get('Vendor1') or row.get('Shamrock'),
                'vendor2': row.get('Vendor2') or row.get('SYSCO'),
                'match_score': self._safe_float(row.get('Match Score')),
                'price_difference': self._safe_float(row.get('Price Diff'))
            }

            if match['product']:
                matching_data['matches'].append(match)
                if match['vendor1']:
                    matching_data['vendors'].add(match['vendor1'])
                if match['vendor2']:
                    matching_data['vendors'].add(match['vendor2'])

        matching_data['vendors'] = list(matching_data['vendors'])
        logger.info(f"Imported {len(matching_data['matches'])} vendor matches")
        return matching_data

    def get_sheet_names(self, filename: str) -> Optional[List[str]]:
        """
        Get list of sheet names in Excel file

        Args:
            filename: Excel filename

        Returns:
            List of sheet names
        """
        file_path = self.data_directory / filename

        if not file_path.exists():
            logger.error(f"File not found: {file_path}")
            return None

        try:
            workbook = load_workbook(file_path, read_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
            return sheet_names
        except Exception as e:
            logger.error(f"Error reading {filename}: {e}")
            return None

    def export_to_excel(self, data: List[Dict], output_filename: str,
                       sheet_name: str = 'Sheet1') -> bool:
        """
        Export data to Excel file

        Args:
            data: List of dictionaries to export
            output_filename: Output Excel filename
            sheet_name: Name of the sheet to create

        Returns:
            True if successful
        """
        if not data:
            logger.error("No data to export")
            return False

        output_path = self.data_directory / output_filename

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = sheet_name

            # Write headers
            headers = list(data[0].keys())
            for col, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=col, value=header)

            # Write data
            for row_idx, row_data in enumerate(data, start=2):
                for col_idx, header in enumerate(headers, start=1):
                    sheet.cell(row=row_idx, column=col_idx, value=row_data.get(header))

            workbook.save(output_path)
            logger.info(f"Exported {len(data)} rows to {output_filename}")
            return True

        except Exception as e:
            logger.error(f"Error exporting to {output_filename}: {e}")
            return False

    def _safe_float(self, value: Any, default: float = 0.0) -> float:
        """Safely convert value to float"""
        try:
            return float(value) if value is not None else default
        except (ValueError, TypeError):
            return default

    def _safe_int(self, value: Any, default: int = 0) -> int:
        """Safely convert value to int"""
        try:
            return int(value) if value is not None else default
        except (ValueError, TypeError):
            return default

    def get_summary(self) -> Dict:
        """
        Get summary of imported data

        Returns:
            Summary dictionary
        """
        return {
            'files_imported': len(set(k.split(':')[0] for k in self.imported_data.keys())),
            'sheets_imported': len(self.imported_data),
            'total_rows': sum(len(data) for data in self.imported_data.values()),
            'files': list(set(k.split(':')[0] for k in self.imported_data.keys()))
        }


if __name__ == "__main__":
    # Test the importer
    importer = ExcelImporter()

    print("Available Excel files:")
    excel_files = importer.list_excel_files()
    for f in excel_files:
        print(f"  - {f}")

    print("\nImporter ready to use!")
